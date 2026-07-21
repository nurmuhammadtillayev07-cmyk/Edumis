"""
Excel / PDF export va Excel import.
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

from .db import WEEKDAYS


def _build_grid(db, class_id):
    """Sinf uchun {slot_order: {day: lesson_row}} shaklidagi grid tuzadi."""
    slots = db.execute("SELECT * FROM tt_slots ORDER BY slot_order").fetchall()
    lessons = db.execute(
        """
        SELECT l.*, s.name AS subject_name, t.full_name AS teacher_name, r.name AS room_name
        FROM tt_lessons l
        JOIN tt_subjects s ON s.id = l.subject_id
        JOIN tt_teachers t ON t.id = l.teacher_id
        JOIN tt_rooms r ON r.id = l.room_id
        WHERE l.class_id = ?
        """,
        (class_id,),
    ).fetchall()

    grid = {s["slot_order"]: {} for s in slots}
    for l in lessons:
        grid[l["slot_order"]][l["day_of_week"]] = l
    return slots, grid


def export_excel(db, class_id, class_name):
    slots, grid = _build_grid(db, class_id)

    wb = Workbook()
    ws = wb.active
    ws.title = class_name[:31]

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value="Vaqt").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for col, (day_num, day_name) in enumerate(WEEKDAYS, start=2):
        c = ws.cell(row=1, column=col, value=day_name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for row_idx, slot in enumerate(slots, start=2):
        ws.cell(row=row_idx, column=1, value=f"{slot['start_time']}–{slot['end_time']}").border = border
        for col, (day_num, day_name) in enumerate(WEEKDAYS, start=2):
            lesson = grid[slot["slot_order"]].get(day_num)
            cell = ws.cell(row=row_idx, column=col)
            if lesson:
                cell.value = (
                    f"{lesson['subject_name']}\n{lesson['teacher_name']}\n{lesson['room_name']}"
                )
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = border

    for col in range(1, len(WEEKDAYS) + 2):
        ws.column_dimensions[chr(64 + col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_pdf(db, class_id, class_name):
    slots, grid = _build_grid(db, class_id)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1 * cm, rightMargin=1 * cm)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(f"<b>{class_name} — Haftalik dars jadvali</b>", styles["Title"]),
        Spacer(1, 0.5 * cm),
    ]

    header = ["Vaqt"] + [name for _, name in WEEKDAYS]
    data = [header]
    for slot in slots:
        row = [f"{slot['start_time']}\n{slot['end_time']}"]
        for day_num, _ in WEEKDAYS:
            lesson = grid[slot["slot_order"]].get(day_num)
            if lesson:
                row.append(f"{lesson['subject_name']}\n{lesson['teacher_name']}\n{lesson['room_name']}")
            else:
                row.append("")
        data.append(row)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf


def import_excel(db, class_id, file_stream):
    """
    Excel fayldan jadval import qilish. Format export_excel bilan bir xil:
    1-ustun vaqt, keyingi 6 ustun Dushanba..Shanba, har katak
    "Fan\\nO'qituvchi\\nXona" formatida.

    Fan/o'qituvchi/xona nomi bazada topilmasa, o'sha qator o'tkazib
    yuboriladi va xato ro'yxatiga qo'shiladi (import to'liq to'xtamaydi).
    """
    wb = load_workbook(file_stream)
    ws = wb.active

    slots = db.execute("SELECT * FROM tt_slots ORDER BY slot_order").fetchall()
    slot_orders = [s["slot_order"] for s in slots]

    imported, errors = 0, []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
        if row_idx >= len(slot_orders):
            break
        slot_order = slot_orders[row_idx]

        for col_idx, day_num in enumerate([d for d, _ in WEEKDAYS], start=1):
            cell_value = row[col_idx] if col_idx < len(row) else None
            if not cell_value:
                continue
            parts = [p.strip() for p in str(cell_value).split("\n") if p.strip()]
            if len(parts) < 3:
                errors.append(f"Qator {row_idx+2}, ustun {col_idx+1}: format noto'g'ri")
                continue
            subject_name, teacher_name, room_name = parts[0], parts[1], parts[2]

            subject = db.execute("SELECT id FROM tt_subjects WHERE name = ?", (subject_name,)).fetchone()
            teacher = db.execute("SELECT id FROM tt_teachers WHERE full_name = ?", (teacher_name,)).fetchone()
            room = db.execute("SELECT id FROM tt_rooms WHERE name = ?", (room_name,)).fetchone()

            if not (subject and teacher and room):
                errors.append(f"Qator {row_idx+2}: '{subject_name}/{teacher_name}/{room_name}' topilmadi")
                continue

            db.execute(
                """
                INSERT INTO tt_lessons (class_id, day_of_week, slot_order, subject_id, teacher_id, room_id)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(class_id, day_of_week, slot_order)
                DO UPDATE SET subject_id=excluded.subject_id, teacher_id=excluded.teacher_id,
                              room_id=excluded.room_id, updated_at=CURRENT_TIMESTAMP
                """,
                (class_id, day_num, slot_order, subject["id"], teacher["id"], room["id"]),
            )
            imported += 1

    db.commit()
    return imported, errors
